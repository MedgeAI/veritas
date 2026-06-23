"""Tests for paperconan adapter."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from engine.static_audit.adapters import paperconan_adapter
from engine.static_audit.adapters.paperconan_adapter import (
    PaperconanAdapterError,
    run_paperconan_scan,
)
from engine.tools.registry import (
    PAPERCONAN_NUMERIC_FORENSICS_TOOL_ID,
    TOOLS,
    coerce_tool_params,
)


def test_paperconan_tool_registered() -> None:
    """Verify paperconan tool is registered in the Tool Registry."""
    assert PAPERCONAN_NUMERIC_FORENSICS_TOOL_ID in TOOLS
    tool = TOOLS[PAPERCONAN_NUMERIC_FORENSICS_TOOL_ID]
    assert tool.tool_id == "paperconan.numeric_forensics"
    assert tool.source == "third_party/paperconan"
    assert tool.agent_selectable is True
    assert tool.deterministic is True
    assert "numeric/paperconan_scan.json" in tool.output_artifacts


def test_paperconan_coerce_params_valid() -> None:
    """Verify paperconan params are coerced correctly."""
    # Default profile
    params = coerce_tool_params(PAPERCONAN_NUMERIC_FORENSICS_TOOL_ID, {})
    assert params == {"profile": "review"}

    # Explicit valid profiles
    for profile in ("review", "forensic", "triage"):
        params = coerce_tool_params(
            PAPERCONAN_NUMERIC_FORENSICS_TOOL_ID, {"profile": profile}
        )
        assert params == {"profile": profile}


def test_paperconan_coerce_params_invalid() -> None:
    """Verify paperconan params reject invalid profile."""
    with pytest.raises(ValueError, match="profile must be one of"):
        coerce_tool_params(PAPERCONAN_NUMERIC_FORENSICS_TOOL_ID, {"profile": "invalid"})


def test_paperconan_adapter_missing_source_data_dir(tmp_path: Path) -> None:
    """Verify adapter raises error when source_data_dir does not exist."""
    nonexistent = tmp_path / "nonexistent"
    output_dir = tmp_path / "output"

    with pytest.raises(PaperconanAdapterError, match="does not exist"):
        run_paperconan_scan(nonexistent, output_dir)


def test_paperconan_adapter_empty_source_data_dir(tmp_path: Path) -> None:
    """Verify adapter handles empty source data directory gracefully."""
    source_data_dir = tmp_path / "source_data"
    source_data_dir.mkdir()
    output_dir = tmp_path / "output"

    result = run_paperconan_scan(source_data_dir, output_dir)

    # Should return "no_data" status, not raise an error
    assert result["status"] == "no_data"
    assert (
        "no .xlsx" in result["error"].lower()
        or "no supported files" in result["error"].lower()
    )
    assert result["findings_summary"]["total"] == 0
    assert result["artifact_path"] is not None

    # Verify error artifact was written
    artifact_path = Path(result["artifact_path"])
    assert artifact_path.exists()
    artifact = json.loads(artifact_path.read_text())
    assert artifact["status"] == "no_data"


def test_paperconan_adapter_with_synthetic_data(tmp_path: Path) -> None:
    """Verify adapter runs successfully on synthetic data with known patterns.

    This test creates a minimal xlsx file with a known fabrication pattern
    (identical columns) and verifies the adapter detects it.
    """
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    # Create synthetic source data with identical columns (a known fabrication pattern)
    source_data_dir = tmp_path / "source_data"
    source_data_dir.mkdir()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write header
    ws["A1"] = "Sample"
    ws["B1"] = "Control"
    ws["C1"] = "Treatment"
    ws["D1"] = "Duplicate_of_Control"

    # Write data: Control and Duplicate_of_Control are identical (fabrication pattern)
    for i in range(2, 12):
        ws[f"A{i}"] = f"Sample_{i - 1}"
        ws[f"B{i}"] = 10.0 + i  # Control: 11, 12, 13, ...
        ws[f"C{i}"] = 15.0 + i  # Treatment: 16, 17, 18, ...
        ws[f"D{i}"] = 10.0 + i  # Duplicate of Control (identical)

    xlsx_path = source_data_dir / "synthetic_data.xlsx"
    wb.save(xlsx_path)

    output_dir = tmp_path / "output"
    result = run_paperconan_scan(source_data_dir, output_dir)

    # Verify scan succeeded
    assert result["status"] == "success"
    assert result["tool"] == "paperconan"
    assert result["tool_version"] != "unknown"

    # Verify findings summary
    summary = result["findings_summary"]
    assert summary["total"] > 0, "Expected at least one finding (identical columns)"
    assert (
        "identical_column" in summary["by_kind"]
        or "constant_offset" in summary["by_kind"]
    ), f"Expected identical_column or constant_offset finding, got {summary['by_kind']}"

    # Verify artifact was written
    artifact_path = Path(result["artifact_path"])
    assert artifact_path.exists()
    artifact = json.loads(artifact_path.read_text())
    assert artifact["status"] == "success"
    assert "scan_result" in artifact

    # Verify scan_result contains the expected structure
    scan_result = artifact["scan_result"]
    assert "tool" in scan_result
    assert scan_result["tool"] == "paperconan"
    assert "relations_blocks" in scan_result
    assert len(scan_result["relations_blocks"]) > 0
    assert not (output_dir / "scan.json").exists()
    assert artifact["artifact_policy"]["upstream_scan_json"] == "disabled"
    assert not _has_key(artifact, "evidence")


def test_paperconan_adapter_strips_bulky_upstream_fields(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source_data_dir = tmp_path / "source_data"
    source_data_dir.mkdir()
    (source_data_dir / "data.csv").write_text("a,b\n1,1\n2,2\n", encoding="utf-8")
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "scan.json").write_text('{"legacy": true}', encoding="utf-8")

    def fake_scan_dir(**_: object) -> dict:
        return {
            "tool": "paperconan",
            "tool_version": "test",
            "relations_blocks": [
                {
                    "file": "data.csv",
                    "sheet": "data",
                    "block": {"rows": "1-2", "cols": "1-2"},
                    "relations": [
                        {
                            "kind": "identical_column",
                            "severity": "high",
                            "evidence": {"rows": [{"values": [1, 2, 3]}]},
                            "raw_values": [1, 2, 3],
                            "base64": "x" * 1000,
                        }
                    ],
                    "progressions": [],
                    "equal_pairs": [],
                    "within_col": [],
                    "identical_after_rounding": [],
                    "grim": [],
                }
            ],
            "cross_sheet_findings": [],
            "digit_distribution": [],
            "decimal_endings": [],
        }

    def fake_load_paperconan() -> tuple[object, type[Exception]]:
        return fake_scan_dir, ValueError

    monkeypatch.setattr(paperconan_adapter, "_load_paperconan", fake_load_paperconan)

    result = run_paperconan_scan(source_data_dir, output_dir)
    artifact = json.loads(Path(result["artifact_path"]).read_text())

    assert result["findings_summary"]["total"] == 1
    assert not (output_dir / "scan.json").exists()
    assert not _has_key(artifact, "evidence")
    assert not _has_key(artifact, "raw_values")
    assert not _has_key(artifact, "base64")


def _has_key(value: object, key: str) -> bool:
    if isinstance(value, dict):
        return key in value or any(_has_key(item, key) for item in value.values())
    if isinstance(value, list):
        return any(_has_key(item, key) for item in value)
    return False
