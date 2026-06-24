"""Tests for source_data detection modules.

Merged from: test_source_data_{findings,pair_forensics,cross_sheet,tool_wrappers}.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from xml.sax.saxutils import escape
import json
import os
import pytest
import subprocess
import zipfile

from engine.static_audit.tools.source_data_cross_sheet import (
    CrossSheetFinding,
    SheetColumn,
    extract_numeric_columns,
    find_cross_sheet_duplicates,
    main,
    run_cross_sheet_detection,
)
from engine.static_audit.tools.source_data_findings import (
    SheetVectors,
    duplicate_column_findings,
    fixed_relationship_findings,
    relationship_record,
)
from engine.static_audit.tools.source_data_pair_forensics import (
    PairForensicsParams,
    analyze_xlsx_root,
    cluster_pair_forensics_findings,
    duplicate_row_vector_findings,
    paired_ratio_reuse_findings,
    pair_forensics_review_tasks,
)


# ===========================================================================
# test_source_data_findings.py
# ===========================================================================


@dataclass
class MockSheetVectors:
    """Mock SheetVectors for testing."""

    workbook: str
    sheet: str
    numeric_columns: dict[int, list[Decimal]]
    text_columns: dict[int, list[tuple[int, str]]] = field(default_factory=dict)
    formulas: dict[int, list[tuple[int, str]]] = field(default_factory=dict)


def test_pattern_strength_complete_35_of_35():
    """Test pattern_strength='complete' when 35/35 rows have fixed_difference=0.3 (FD-0001 scenario)."""
    # Create mock data: 35 rows with fixed_difference=0.3
    left_col_data = [Decimal("0")] * 40  # 40 rows total
    right_col_data = [Decimal("0")] * 40

    # Fill rows 5-39 (35 rows) with fixed difference = 0.3
    for i in range(5, 40):
        left_col_data[i] = Decimal("1.0") + Decimal(str((i - 5) * 0.1))
        right_col_data[i] = left_col_data[i] - Decimal("0.3")

    sheet = MockSheetVectors(
        workbook="test_workbook.xlsx",
        sheet="Source Data Fig.4",
        numeric_columns={0: left_col_data, 1: right_col_data},
    )

    result = relationship_record(
        sheet=sheet,
        relationship="fixed_difference",
        left_col=0,
        right_col=1,
        value="0.3",
        support_rows=35,
        overlap_rows=35,
        rows=list(range(5, 40)),
        formula_cols=set(),
    )

    assert result["pattern_strength"] == "complete"
    assert (
        result["pattern_strength_reason"]
        == "fixed_difference=0.3 covers 35/35 overlapping rows"
    )
    assert result["support_rate"] == 1.0
    assert result["risk_level"] == "medium"  # 35 < 100, so medium


def test_pattern_strength_strong_80_percent():
    """Test pattern_strength='strong' when 80% of rows have fixed relationship."""
    left_col_data = [Decimal("0")] * 100
    right_col_data = [Decimal("0")] * 100

    # 80 rows with fixed difference, 20 rows without
    for i in range(100):
        left_col_data[i] = Decimal("1.0") + Decimal(str(i * 0.1))
        if i < 80:
            right_col_data[i] = left_col_data[i] - Decimal("0.5")
        else:
            right_col_data[i] = left_col_data[i] - Decimal("999")  # Different pattern

    sheet = MockSheetVectors(
        workbook="test_workbook.xlsx",
        sheet="Test Sheet",
        numeric_columns={0: left_col_data, 1: right_col_data},
    )

    result = relationship_record(
        sheet=sheet,
        relationship="fixed_difference",
        left_col=0,
        right_col=1,
        value="0.5",
        support_rows=80,
        overlap_rows=100,
        rows=list(range(100)),
        formula_cols=set(),
    )

    assert result["pattern_strength"] == "strong"
    assert (
        result["pattern_strength_reason"]
        == "fixed_difference=0.5 covers 80/100 overlapping rows"
    )
    assert result["support_rate"] == 0.8
    assert result["risk_level"] == "medium"  # 80 < 100, so medium


def test_pattern_strength_high_when_support_rows_ge_100():
    """Test risk_level='high' when support_rows >= 100, regardless of pattern_strength."""
    left_col_data = [Decimal("0")] * 150
    right_col_data = [Decimal("0")] * 150

    # 100 rows with fixed difference
    for i in range(150):
        left_col_data[i] = Decimal("1.0") + Decimal(str(i * 0.1))
        if i < 100:
            right_col_data[i] = left_col_data[i] - Decimal("0.2")
        else:
            right_col_data[i] = left_col_data[i] - Decimal("999")

    sheet = MockSheetVectors(
        workbook="test_workbook.xlsx",
        sheet="Test Sheet",
        numeric_columns={0: left_col_data, 1: right_col_data},
    )

    result = relationship_record(
        sheet=sheet,
        relationship="fixed_difference",
        left_col=0,
        right_col=1,
        value="0.2",
        support_rows=100,
        overlap_rows=150,
        rows=list(range(150)),
        formula_cols=set(),
    )

    assert result["pattern_strength"] == "moderate"  # 100/150 = 66.7% -> moderate
    assert result["risk_level"] == "high"  # 100 >= 100, so high
    assert result["support_rate"] == round(100 / 150, 4)


def test_pattern_strength_with_formula_column():
    """Test pattern_strength when formula column is involved (risk_level='low')."""
    left_col_data = [Decimal("0")] * 50
    right_col_data = [Decimal("0")] * 50

    # 50 rows with fixed difference
    for i in range(50):
        left_col_data[i] = Decimal("1.0") + Decimal(str(i * 0.1))
        right_col_data[i] = left_col_data[i] - Decimal("0.4")

    sheet = MockSheetVectors(
        workbook="test_workbook.xlsx",
        sheet="Test Sheet",
        numeric_columns={0: left_col_data, 1: right_col_data},
    )

    result = relationship_record(
        sheet=sheet,
        relationship="fixed_difference",
        left_col=0,
        right_col=1,
        value="0.4",
        support_rows=50,
        overlap_rows=50,
        rows=list(range(50)),
        formula_cols={1},  # right_col (1) is a formula column
    )

    assert result["pattern_strength"] == "complete"  # 50/50 = 100%
    assert result["risk_level"] == "low"  # formula_involved -> low
    assert result["formula_column_involved"] is True
    assert result["artifact_likelihood"] == "medium"


def test_pattern_strength_moderate_and_weak():
    """Test pattern_strength='moderate' and 'weak' scenarios."""
    # Moderate: 60% coverage
    sheet = MockSheetVectors(
        workbook="test_workbook.xlsx",
        sheet="Test Sheet",
        numeric_columns={0: [Decimal("0")] * 100, 1: [Decimal("0")] * 100},
    )

    result = relationship_record(
        sheet=sheet,
        relationship="fixed_ratio",
        left_col=0,
        right_col=1,
        value="1.5",
        support_rows=60,
        overlap_rows=100,
        rows=list(range(100)),
        formula_cols=set(),
    )

    assert result["pattern_strength"] == "moderate"
    assert result["support_rate"] == 0.6

    # Weak: 30% coverage
    result = relationship_record(
        sheet=sheet,
        relationship="fixed_ratio",
        left_col=0,
        right_col=1,
        value="1.5",
        support_rows=30,
        overlap_rows=100,
        rows=list(range(100)),
        formula_cols=set(),
    )

    assert result["pattern_strength"] == "weak"
    assert result["support_rate"] == 0.3


def test_mean_sum_n_relationship_is_artifact_not_priority_signal() -> None:
    rows = list(range(10, 20))
    sheet = SheetVectors(
        workbook="source.xlsx",
        workbook_path="source.xlsx",
        sheet="Extended Data Fig. 3d",
        sheet_path="xl/worksheets/sheet1.xml",
        numeric_columns={
            4: {row: Decimal("100") for row in rows},
            5: {row: Decimal(row) for row in rows},
            7: {row: Decimal(row * 100) for row in rows},
        },
        text_columns={
            4: [(6, "N total")],
            5: [(6, "Mean"), (9, "Output voltage (V)")],
            7: [(6, "Sum"), (9, "Voltage (V)")],
        },
        formulas_by_column={},
        cell_count=30,
        numeric_cell_count=30,
    )

    findings = fixed_relationship_findings(
        sheet,
        min_overlap=8,
        min_support=0.98,
        limit=20,
    )
    finding = next(item for item in findings if item["category"] == "fixed_ratio")

    assert finding["risk_level"] == "low"
    assert finding["artifact_likelihood"] == "high"
    assert finding["artifact_reason"] == "Mean/Sum/N summary-statistics relationship"
    assert finding["pressure_test_result"] == "likely_summary_statistic_derivation"


def test_zero_inflated_duplicate_columns_are_downgraded() -> None:
    rows = list(range(1, 51))
    sheet = SheetVectors(
        workbook="source.xlsx",
        workbook_path="source.xlsx",
        sheet="Tumor free animals",
        sheet_path="xl/worksheets/sheet1.xml",
        numeric_columns={
            2: {row: Decimal("0") for row in rows},
            3: {row: Decimal("0") for row in rows},
        },
        text_columns={
            2: [(1, "Group A response")],
            3: [(1, "Group B response")],
        },
        formulas_by_column={},
        cell_count=100,
        numeric_cell_count=100,
    )
    sheet.numeric_columns[2][50] = Decimal("1")
    sheet.numeric_columns[3][50] = Decimal("2")

    findings = duplicate_column_findings(
        sheet,
        min_overlap=20,
        min_support=0.98,
        limit=20,
    )

    assert len(findings) == 1
    assert findings[0]["risk_level"] == "low"
    assert findings[0]["artifact_likelihood"] == "high"
    assert findings[0]["pressure_test_result"] == "likely_zero_inflated_matrix_artifact"


def test_mean_sum_labels_without_integer_n_relationship_are_not_downgraded() -> None:
    rows = list(range(10, 20))
    sheet = SheetVectors(
        workbook="source.xlsx",
        workbook_path="source.xlsx",
        sheet="Independent endpoint summary",
        sheet_path="xl/worksheets/sheet1.xml",
        numeric_columns={
            5: {row: Decimal(row) for row in rows},
            7: {row: Decimal(row) * Decimal("2.5") for row in rows},
        },
        text_columns={
            5: [(6, "Mean response")],
            7: [(6, "Sum response")],
        },
        formulas_by_column={},
        cell_count=20,
        numeric_cell_count=20,
    )

    findings = fixed_relationship_findings(
        sheet,
        min_overlap=8,
        min_support=0.98,
        limit=20,
    )
    finding = next(item for item in findings if item["category"] == "fixed_ratio")

    assert finding["risk_level"] == "medium"
    assert finding["artifact_likelihood"] == "unknown"
    assert finding["pressure_test_result"] == "needs_semantics_and_formula_review"


def test_duplicate_columns_below_zero_inflation_threshold_are_not_downgraded() -> None:
    rows = list(range(1, 21))
    sheet = SheetVectors(
        workbook="source.xlsx",
        workbook_path="source.xlsx",
        sheet="Endpoint matrix",
        sheet_path="xl/worksheets/sheet1.xml",
        numeric_columns={
            2: {},
            3: {},
        },
        text_columns={
            2: [(1, "Endpoint A")],
            3: [(1, "Endpoint B")],
        },
        formulas_by_column={},
        cell_count=40,
        numeric_cell_count=40,
    )
    for row in rows:
        value = Decimal("0") if row <= 13 else Decimal(row)
        sheet.numeric_columns[2][row] = value
        sheet.numeric_columns[3][row] = value

    findings = duplicate_column_findings(
        sheet,
        min_overlap=20,
        min_support=0.98,
        limit=20,
    )

    assert len(findings) == 1
    assert findings[0]["risk_level"] == "medium"
    assert findings[0]["artifact_likelihood"] == "unknown"
    assert findings[0]["pressure_test_result"] == "needs_column_semantics_review"


# ===========================================================================
# test_source_data_pair_forensics.py
# ===========================================================================


def write_minimal_xlsx(path: Path, rows: list[list[float | int | None]]) -> None:
    sheet_rows = []
    for row_index, values in enumerate(rows, start=1):
        cells = []
        for col_index, value in enumerate(values, start=1):
            if value is None:
                continue
            col = chr(64 + col_index)
            cells.append(f'<c r="{col}{row_index}"><v>{value}</v></c>')
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            "</Types>",
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Fig.1a" sheetId="1" r:id="rId1"/></sheets>'
            "</workbook>",
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            "</Relationships>",
        )
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def write_mixed_xlsx(path: Path, rows: list[list[float | int | str | None]]) -> None:
    sheet_rows = []
    for row_index, values in enumerate(rows, start=1):
        cells = []
        for col_index, value in enumerate(values, start=1):
            if value is None:
                continue
            col = chr(64 + col_index)
            if isinstance(value, str):
                cells.append(
                    f'<c r="{col}{row_index}" t="inlineStr"><is><t>{escape(value)}</t></is></c>'
                )
            else:
                cells.append(f'<c r="{col}{row_index}"><v>{value}</v></c>')
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            "</Types>",
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Fig.1a" sheetId="1" r:id="rId1"/></sheets>'
            "</workbook>",
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            "</Relationships>",
        )
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def test_pair_forensics_detects_row_offset_ratio_and_scalar_patterns(tmp_path) -> None:
    # Rows 1-4 and 5-8 are not paper-specific. They model the generic pattern:
    # a second block reuses the first block's paired ratios and scalar-multiplies
    # the underlying columns at a fixed row offset.
    write_minimal_xlsx(
        tmp_path / "source.xlsx",
        [
            [1, 2],
            [2, 6],
            [3, 12],
            [4, 20],
            [10, 20],
            [20, 60],
            [30, 120],
            [40, 200],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [7, 8],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [None, None],
            [7, 8],
        ],
    )

    result = analyze_xlsx_root(
        tmp_path,
        PairForensicsParams(min_pairs=4, min_support=1.0, max_offset=8),
    )
    categories = {item["category"] for item in result["findings"]}

    assert "paired_ratio_reuse" in categories
    assert "row_offset_scalar_multiple" in categories
    assert "duplicate_row_vector" in categories
    assert result["summary"]["priority_findings"] >= 3
    assert result["summary"]["finding_clusters"] >= 1
    assert result["summary"]["review_tasks"] >= 1
    assert result["finding_clusters"]
    assert result["review_tasks"]


def test_pair_forensics_detects_long_format_paired_ratio_reuse(tmp_path) -> None:
    # Long-format source data is common in biomedical papers: one column stores
    # pair/sample id and another stores the measurement for two consecutive rows.
    write_minimal_xlsx(
        tmp_path / "long_format.xlsx",
        [
            [None, 1, 1],
            [None, 1, 2],
            [None, 2, 2],
            [None, 2, 6],
            [None, 3, 3],
            [None, 3, 12],
            [None, 4, 4],
            [None, 4, 20],
            [None, 5, 10],
            [None, 5, 20],
            [None, 6, 20],
            [None, 6, 60],
            [None, 7, 30],
            [None, 7, 120],
            [None, 8, 40],
            [None, 8, 200],
        ],
    )

    result = analyze_xlsx_root(
        tmp_path,
        PairForensicsParams(min_pairs=4, min_support=1.0, max_offset=4),
    )
    categories = {item["category"] for item in result["findings"]}

    assert "long_format_paired_ratio_reuse" in categories


def test_pair_forensics_detects_cross_block_paired_diff_too_narrow(tmp_path) -> None:
    write_mixed_xlsx(
        tmp_path / "cross_block.xlsx",
        [
            [10.00, 100.00],
            [20.00, 200.00],
            [30.00, 300.00],
            [40.00, 400.00],
            ["Treatment B", "Treatment B"],
            [10.01, 100.01],
            [20.01, 200.01],
            [30.01, 300.01],
            [40.01, 400.01],
        ],
    )

    result = analyze_xlsx_root(
        tmp_path,
        PairForensicsParams(min_pairs=3, max_findings_per_category=20),
    )

    findings = [
        item
        for item in result["findings"]
        if item["category"] == "cross_block_paired_diff_too_narrow"
    ]
    assert findings
    assert result["summary"]["cross_block_paired_diff_too_narrow_findings"] == len(
        findings
    )
    assert any(item["finding_id"].startswith("CBD-") for item in findings)
    assert "cross_block_paired_diff_too_narrow" in {
        item["category"] for item in result["priority_findings"]
    }
    assert any(
        task["category"] == "cross_block_paired_diff_too_narrow"
        for task in result["review_tasks"]
    )


def test_pair_forensics_cross_block_requires_real_separator_and_narrow_diffs(
    tmp_path,
) -> None:
    write_mixed_xlsx(
        tmp_path / "single_cell_separator.xlsx",
        [
            [10.00, 100.00],
            [20.00, 200.00],
            [30.00, 300.00],
            [40.00, 400.00],
            ["Treatment B", None],
            [10.01, 100.01],
            [20.01, 200.01],
            [30.01, 300.01],
            [40.01, 400.01],
        ],
    )
    write_mixed_xlsx(
        tmp_path / "wide_diffs.xlsx",
        [
            [10.00, 100.00],
            [20.00, 200.00],
            [30.00, 300.00],
            [40.00, 400.00],
            ["Treatment B", "Treatment B"],
            [15.00, 80.00],
            [18.00, 260.00],
            [45.00, 330.00],
            [62.00, 390.00],
        ],
    )

    result = analyze_xlsx_root(
        tmp_path,
        PairForensicsParams(min_pairs=3, max_findings_per_category=20),
    )

    assert "cross_block_paired_diff_too_narrow" not in {
        item["category"] for item in result["findings"]
    }


def test_pair_forensics_cli_outputs_empty_summary(tmp_path) -> None:
    output = tmp_path / "pair_forensics.json"
    result = analyze_xlsx_root(tmp_path, PairForensicsParams())
    output.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["summary"]["workbook_count"] == 0
    assert data["summary"]["findings"] == 0
    assert data["summary"]["finding_clusters"] == 0
    assert data["summary"]["review_tasks"] == 0


def test_pair_forensics_clusters_repeated_same_sheet_offset_findings() -> None:
    findings = [
        {
            "finding_id": f"PRR-{idx:04d}",
            "category": "paired_ratio_reuse",
            "risk_level": "high",
            "confidence": "high",
            "workbook": "source.xlsx",
            "sheet": "Fig1",
            "row_offset": 8,
            "column_pair": [f"A{idx}", f"B{idx}"],
            "matched_pairs": 12,
            "overlap_pairs": 12,
            "support_rate": 1.0,
            "benign_explanations": ["可能是合法派生。"],
            "next_steps": ["核对原始记录。"],
        }
        for idx in range(1, 6)
    ]

    clusters = cluster_pair_forensics_findings(findings)
    tasks = pair_forensics_review_tasks(clusters)

    assert len(clusters) == 1
    assert clusters[0]["finding_count"] == 5
    assert clusters[0]["pattern_signature"] == "offset=8"
    assert len(clusters[0]["representative_finding_ids"]) == 5
    assert len(tasks) == 1
    assert tasks[0]["cluster_id"] == clusters[0]["cluster_id"]
    assert tasks[0]["cluster_count"] == 1
    assert tasks[0]["finding_count"] == 5


def test_pair_forensics_downgrades_mean_sum_ratio_reuse() -> None:
    rows = list(range(10, 20))
    sheet = SheetVectors(
        workbook="source.xlsx",
        workbook_path="source.xlsx",
        sheet="Extended Data Fig. 3d",
        sheet_path="xl/worksheets/sheet1.xml",
        numeric_columns={
            4: {row: Decimal("100") for row in rows},
            5: {row: Decimal(row) for row in rows},
            7: {row: Decimal(row * 100) for row in rows},
        },
        text_columns={
            4: [(6, "N total")],
            5: [(6, "Mean"), (9, "Output voltage (V)")],
            7: [(6, "Sum"), (9, "Voltage (V)")],
        },
        formulas_by_column={},
        cell_count=30,
        numeric_cell_count=30,
    )

    findings = paired_ratio_reuse_findings(
        sheet,
        PairForensicsParams(min_pairs=4, min_support=1.0, max_offset=4),
    )

    assert findings
    assert {finding["risk_level"] for finding in findings} == {"low"}
    assert {finding["artifact_likelihood"] for finding in findings} == {"high"}
    assert {finding["pressure_test_result"] for finding in findings} == {
        "likely_summary_statistic_derivation"
    }


def test_duplicate_row_vector_downgrades_zero_inflated_matrix() -> None:
    rows = list(range(1, 41))
    numeric_columns: dict[int, dict[int, Decimal]] = {}
    for col in range(2, 18):
        numeric_columns[col] = {}
        for row in rows:
            if row <= 35:
                numeric_columns[col][row] = Decimal("0")
            else:
                numeric_columns[col][row] = Decimal(col * 100 + row)

    sheet = SheetVectors(
        workbook="source.xlsx",
        workbook_path="source.xlsx",
        sheet="Zero inflated matrix",
        sheet_path="xl/worksheets/sheet1.xml",
        numeric_columns=numeric_columns,
        text_columns={col: [(1, f"Endpoint {col}")] for col in numeric_columns},
        formulas_by_column={},
        cell_count=640,
        numeric_cell_count=640,
    )

    findings = duplicate_row_vector_findings(
        sheet,
        PairForensicsParams(min_duplicate_row_width=2),
    )

    finding = next(item for item in findings if item["duplicate_row_count"] == 35)
    assert finding["risk_level"] == "medium"
    assert finding["artifact_likelihood"] == "high"
    assert finding["pressure_test_result"] == "likely_zero_inflated_matrix_artifact"


# ===========================================================================
# test_source_data_cross_sheet.py
# ===========================================================================


def test_cross_sheet_finding_to_dict() -> None:
    """Test CrossSheetFinding serialization."""
    finding = CrossSheetFinding(
        finding_id="CSD-0001",
        workbook_1="wb1.xlsx",
        sheet_1="Sheet1",
        column_1="A",
        column_1_label="Values",
        workbook_2="wb2.xlsx",
        sheet_2="Sheet2",
        column_2="B",
        column_2_label="Numbers",
        overlap_rows=20,
        equal_rows=18,
        support_rate=0.9,
    )

    data = finding.to_dict()

    assert data["finding_id"] == "CSD-0001"
    assert data["category"] == "cross_sheet_duplicate_columns"
    assert data["issue_category"] == "consistency"
    assert data["overlap_rows"] == 20
    assert data["equal_rows"] == 18
    assert data["support_rate"] == 0.9


def test_extract_numeric_columns_with_openpyxl(tmp_path: Path) -> None:
    """Test numeric column extraction from XLSX file."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add header
    ws.append(["ID", "Value1", "Value2", "Text"])

    # Add data rows
    for i in range(1, 11):
        ws.append([i, float(i) * 1.5, float(i) * 2.0, f"row_{i}"])

    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)

    columns = extract_numeric_columns(xlsx_path, "Sheet1")

    # Should extract ID, Value1, Value2 (3 numeric columns)
    assert len(columns) == 3

    # Check that Value1 column has correct values
    value1_col = [c for c in columns if c.column_label == "Value1"][0]
    assert len(value1_col.values) == 10
    assert value1_col.values[0] == 1.5
    assert value1_col.values[9] == 15.0


def test_find_cross_sheet_duplicates_identifies_duplicates(tmp_path: Path) -> None:
    """Test that cross-sheet duplicate detection finds duplicates."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    # Create two XLSX files with overlapping numeric columns
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Value"])
    for i in range(1, 21):
        ws1.append([i, float(i) * 2.0])
    wb1.save(tmp_path / "file1.xlsx")

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet2"
    ws2.append(["ID", "Value"])
    # Same values as file1 for first 15 rows
    for i in range(1, 16):
        ws2.append([i, float(i) * 2.0])
    # Different values for remaining rows
    for i in range(16, 21):
        ws2.append([i, float(i) * 3.0])
    wb2.save(tmp_path / "file2.xlsx")

    findings = find_cross_sheet_duplicates(
        tmp_path,
        min_overlap=10,
        min_support_rate=0.7,
        max_findings=10,
    )

    # Should find at least one duplicate (ID columns match perfectly)
    assert len(findings) >= 1

    # Check that Value columns are detected as duplicates (15/20 = 0.75 support)
    value_findings = [
        f
        for f in findings
        if "Value" in f.column_1_label or "Value" in f.column_2_label
    ]
    if value_findings:
        # Support rate should be around 0.75 (15 matching out of 20 overlapping)
        assert 0.7 <= value_findings[0].support_rate <= 0.8


def test_find_cross_sheet_duplicates_respects_min_overlap(tmp_path: Path) -> None:
    """Test that min_overlap parameter is respected."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    # Create files with only 5 overlapping rows
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(["Value"])
    for i in range(1, 6):
        ws1.append([float(i)])
    wb1.save(tmp_path / "file1.xlsx")

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Value"])
    for i in range(1, 6):
        ws2.append([float(i)])
    wb2.save(tmp_path / "file2.xlsx")

    # With min_overlap=10, should find no duplicates
    findings = find_cross_sheet_duplicates(
        tmp_path,
        min_overlap=10,
        min_support_rate=0.8,
        max_findings=10,
    )
    assert len(findings) == 0

    # With min_overlap=5, should find duplicates
    findings = find_cross_sheet_duplicates(
        tmp_path,
        min_overlap=5,
        min_support_rate=0.8,
        max_findings=10,
    )
    assert len(findings) >= 1


def test_find_cross_sheet_duplicates_respects_min_support_rate(tmp_path: Path) -> None:
    """Test that min_support_rate parameter is respected."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    # Create files with 50% matching values
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(["Value"])
    for i in range(1, 21):
        ws1.append([float(i)])
    wb1.save(tmp_path / "file1.xlsx")

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Value"])
    # First 10 rows match, next 10 don't
    for i in range(1, 11):
        ws2.append([float(i)])
    for i in range(11, 21):
        ws2.append([float(i) * 10.0])  # Different values
    wb2.save(tmp_path / "file2.xlsx")

    # With min_support_rate=0.8, should find no duplicates (only 50% match)
    findings = find_cross_sheet_duplicates(
        tmp_path,
        min_overlap=10,
        min_support_rate=0.8,
        max_findings=10,
    )
    assert len(findings) == 0

    # With min_support_rate=0.4, should find duplicates
    findings = find_cross_sheet_duplicates(
        tmp_path,
        min_overlap=10,
        min_support_rate=0.4,
        max_findings=10,
    )
    assert len(findings) >= 1


def test_run_cross_sheet_detection_output_format(tmp_path: Path) -> None:
    """Test that run_cross_sheet_detection returns correct format."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    # Create a simple XLSX file
    wb = Workbook()
    ws = wb.active
    ws.append(["Value"])
    for i in range(1, 11):
        ws.append([float(i)])
    wb.save(tmp_path / "test.xlsx")

    result = run_cross_sheet_detection(
        tmp_path,
        min_overlap=5,
        min_support_rate=0.8,
        max_findings=10,
    )

    # Check output structure
    assert "findings" in result
    assert "parameters" in result
    assert result["parameters"]["min_overlap"] == 5
    assert result["parameters"]["min_support_rate"] == 0.8
    assert result["parameters"]["max_findings"] == 10

    # findings should be a list
    assert isinstance(result["findings"], list)


def test_run_cross_sheet_detection_with_no_xlsx_files(tmp_path: Path) -> None:
    """Test detection with directory containing no XLSX files."""
    result = run_cross_sheet_detection(
        tmp_path,
        min_overlap=10,
        min_support_rate=0.8,
        max_findings=10,
    )

    assert result["findings"] == []
    assert "parameters" in result


def test_cli_accepts_output_flag_and_parameters(tmp_path: Path) -> None:
    output = tmp_path / "nested" / "source_data_cross_sheet.json"

    exit_code = main(
        [
            str(tmp_path),
            "--output",
            str(output),
            "--min-overlap",
            "5",
            "--min-support",
            "0.7",
            "--max-findings",
            "3",
        ]
    )

    assert exit_code == 0
    assert output.exists()
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["parameters"] == {
        "min_overlap": 5,
        "min_support_rate": 0.7,
        "max_findings": 3,
    }


def test_sheet_column_dataclass() -> None:
    """Test SheetColumn dataclass."""
    col = SheetColumn(
        workbook="test.xlsx",
        sheet="Sheet1",
        column="A",
        column_label="Value",
        values=[10.0, 20.0, 30.0],
        row_indices=[1, 2, 3],
    )

    assert col.workbook == "test.xlsx"
    assert col.sheet == "Sheet1"
    assert col.column == "A"
    assert col.column_label == "Value"
    assert len(col.values) == 3
    assert len(col.row_indices) == 3


# ===========================================================================
# test_source_data_tool_wrappers.py
# ===========================================================================


PROJECT_ROOT = Path(__file__).resolve().parents[2]


def run_script(args: list[str]) -> subprocess.CompletedProcess[str]:
    env = dict(os.environ)
    env["PYTHONPATH"] = "."
    return subprocess.run(
        args,
        cwd=PROJECT_ROOT,
        env=env,
        capture_output=True,
        text=True,
        check=False,
    )


def test_source_data_profile_wrapper_outputs_stable_empty_summary(tmp_path) -> None:
    source_root = tmp_path / "Source Data"
    source_root.mkdir()
    output = tmp_path / "source_data_profile.json"

    result = run_script(
        [
            "python3",
            "scripts/source_data_profile.py",
            str(source_root),
            "--output",
            str(output),
        ]
    )

    assert result.returncode == 0, result.stderr
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["summary"]["workbook_count"] == 0
    assert data["summary"]["sheet_count"] == 0


def test_source_data_findings_wrapper_outputs_stable_empty_summary(tmp_path) -> None:
    source_root = tmp_path / "Source Data"
    source_root.mkdir()
    profile = tmp_path / "source_data_profile.json"
    profile.write_text(
        json.dumps(
            {"summary": {"workbook_count": 0, "sheet_count": 0}, "workbooks": []}
        ),
        encoding="utf-8",
    )
    output = tmp_path / "source_data_findings.json"

    result = run_script(
        [
            "python3",
            "scripts/source_data_findings.py",
            str(source_root),
            "--profile",
            str(profile),
            "--output",
            str(output),
        ]
    )

    assert result.returncode == 0, result.stderr
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["summary"]["workbook_count"] == 0
    assert data["summary"]["priority_findings"] == 0


def test_source_data_pair_forensics_wrapper_outputs_stable_empty_summary(
    tmp_path,
) -> None:
    source_root = tmp_path / "Source Data"
    source_root.mkdir()
    output = tmp_path / "source_data_pair_forensics.json"

    result = run_script(
        [
            "python3",
            "scripts/source_data_pair_forensics.py",
            str(source_root),
            "--output",
            str(output),
        ]
    )

    assert result.returncode == 0, result.stderr
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["summary"]["workbook_count"] == 0
    assert data["summary"]["priority_findings"] == 0
