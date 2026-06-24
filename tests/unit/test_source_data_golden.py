"""
Contract test for paper4 Fig3 golden fixture.

This test documents the ground truth semantic structure of paper4's Fig3 sheet.
The fixture contains:
- 34 experimental groups across 8 compound classes
- Column semantics (CFU raw, survival %, average, AUC)
- 9 cross-group CFU-identical row pairs
- 3 consecutive-CFU-reuse patterns between group pairs
- 1 full-row duplicate pair (Flavanones::7 row 2 & Isoflavones::11 row 2)

This test does NOT test the Agent (that's future work) — it documents ground truth.
"""
import json
import pytest
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


FIXTURE_PATH = Path(__file__).parent.parent / "fixtures" / "source_data" / "paper4_fig3_golden.json"
# XLSX file is gitignored, may only exist in main repo
XLSX_SEARCH_PATHS = [
    Path(__file__).parent.parent.parent.parent / "input" / "paper4" / "41467_2022_28908_MOESM4_ESM.xlsx",
    Path("/mnt/disk1/LZJ/project/veritas/input/paper4/41467_2022_28908_MOESM4_ESM.xlsx"),
]
XLSX_PATH = next((p for p in XLSX_SEARCH_PATHS if p.exists()), XLSX_SEARCH_PATHS[0])


@pytest.fixture
def golden_fixture():
    """Load the golden fixture."""
    assert FIXTURE_PATH.exists(), f"Golden fixture not found: {FIXTURE_PATH}"
    with open(FIXTURE_PATH) as f:
        return json.load(f)


class TestFixtureStructure:
    """Verify the golden fixture is well-formed."""

    def test_metadata_fields(self, golden_fixture):
        """Verify metadata section has all required fields."""
        meta = golden_fixture["metadata"]
        required = [
            "paper", "sheet", "source_file", "description",
            "total_groups", "total_data_rows",
            "cross_group_reuse_count", "full_row_duplicate_count"
        ]
        for field in required:
            assert field in meta, f"Missing metadata field: {field}"

        assert meta["paper"] == "paper4"
        assert meta["sheet"] == "Fig3"
        assert meta["total_groups"] == 34
        assert meta["total_data_rows"] == 163
        assert meta["cross_group_reuse_count"] == 9

    def test_column_semantics(self, golden_fixture):
        """Verify column_semantics section."""
        cs = golden_fixture["column_semantics"]
        assert len(cs) == 4, "Expected 4 column semantic groups"

        # Verify each has required fields
        for entry in cs:
            assert "columns" in entry
            assert "column_indices" in entry
            assert "data_type" in entry
            assert "description" in entry

        # Verify column indices match expected layout
        b_e = next(e for e in cs if e["columns"] == "B-E")
        assert b_e["column_indices"] == [2, 3, 4, 5]
        assert b_e["data_type"] == "numeric"

        f_i = next(e for e in cs if e["columns"] == "F-I")
        assert f_i["column_indices"] == [6, 7, 8, 9]
        assert f_i["data_type"] == "percentage"

        j_m = next(e for e in cs if e["columns"] == "J-M")
        assert j_m["column_indices"] == [10, 11, 12, 13]

        n_q = next(e for e in cs if e["columns"] == "N-Q")
        assert n_q["column_indices"] == [14, 15, 16, 17]

    def test_groups_structure(self, golden_fixture):
        """Verify groups section has all required fields."""
        groups = golden_fixture["groups"]
        assert len(groups) == 34, f"Expected 34 groups, got {len(groups)}"

        required_fields = [
            "name", "section", "compound_class", "variant",
            "row_start", "row_end", "replicate_count", "data_rows"
        ]
        for g in groups:
            for field in required_fields:
                assert field in g, f"Group {g.get('name', 'unknown')} missing field: {field}"

        # Verify compound classes
        compound_classes = {g["compound_class"] for g in groups}
        expected_classes = {
            "Coumarins", "Benzenoids", "Flavanones",
            "Isoflavones", "Flavonol", "Chalcones"
        }
        assert compound_classes == expected_classes, f"Unexpected compound classes: {compound_classes}"

        # Verify total replicate count
        total_replicates = sum(g["replicate_count"] for g in groups)
        assert total_replicates == 163, f"Expected 163 total replicates, got {total_replicates}"

    def test_variant_types(self, golden_fixture):
        """Verify variant types are correctly classified."""
        groups = golden_fixture["groups"]

        # Count variants
        wt_count = sum(1 for g in groups if g["variant"] == "wt")
        fob_count = sum(1 for g in groups if g["variant"] == "fob")
        fob1_count = sum(1 for g in groups if g["variant"] == "fob1")
        mutant_count = sum(1 for g in groups if g["variant"].startswith("mutant_"))

        # Each compound class should have exactly 1 wt (6 groups total)
        assert wt_count == 6

        # fob and fob1 variants
        assert fob_count >= 1, "Expected at least 1 fob variant"
        assert fob1_count >= 1, "Expected at least 1 fob1 variant"

        # Numbered mutant groups
        assert mutant_count > 0, "Expected numbered mutant groups"

    def test_known_reuse_structure(self, golden_fixture):
        """Verify known_reuse section has all required fields."""
        kr = golden_fixture["known_reuse"]
        assert len(kr) == 9, f"Expected 9 known_reuse entries, got {len(kr)}"

        required_fields = [
            "rows_a", "rows_b", "group_a", "group_b",
            "shared_cfu_values", "full_row_match",
            "row_data_a", "row_data_b"
        ]
        for entry in kr:
            for field in required_fields:
                assert field in entry, f"known_reuse entry missing field: {field}"

            # Verify row_data has correct length (B-R = 17 columns)
            assert len(entry["row_data_a"]) == 17, "row_data_a should have 17 values (B-R)"
            assert len(entry["row_data_b"]) == 17, "row_data_b should have 17 values (B-R)"

            # Verify shared_cfu_values has 4 values (B-E)
            assert len(entry["shared_cfu_values"]) == 4, "shared_cfu_values should have 4 values (B-E)"

    def test_expected_findings_structure(self, golden_fixture):
        """Verify expected_findings section."""
        ef = golden_fixture["expected_findings"]
        assert len(ef) == 3, f"Expected 3 expected_findings entries, got {len(ef)}"

        # Verify finding types
        finding_types = {f["type"] for f in ef}
        expected_types = {
            "cross_group_data_reuse",
            "consecutive_cfu_reuse",
            "full_row_duplicate"
        }
        assert finding_types == expected_types, f"Unexpected finding types: {finding_types}"


class TestKnownReuseAccuracy:
    """Verify known_reuse entries match actual xlsx data."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not available")
    def test_cross_group_cfu_values_match(self, golden_fixture):
        """Verify that known_reuse CFU values match actual xlsx data."""
        if not XLSX_PATH.exists():
            pytest.skip(f"XLSX file not found: {XLSX_PATH}")

        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        ws = wb["Fig3"]

        for entry in golden_fixture["known_reuse"]:
            r1 = entry["rows_a"][0]
            r2 = entry["rows_b"][0]
            expected_cfu = entry["shared_cfu_values"]

            # Read CFU values from xlsx (columns B-E = 2-5)
            actual_cfu_r1 = [ws.cell(r1, c).value for c in range(2, 6)]
            actual_cfu_r2 = [ws.cell(r2, c).value for c in range(2, 6)]

            # Both rows should have the same CFU values
            assert actual_cfu_r1 == actual_cfu_r2, \
                f"Rows {r1} and {r2} should have identical CFU values"

            # CFU values should match expected
            assert actual_cfu_r1 == expected_cfu, \
                f"Row {r1} CFU values {actual_cfu_r1} don't match expected {expected_cfu}"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not available")
    def test_full_row_matches(self, golden_fixture):
        """Verify that full_row_match entries actually have identical full rows."""
        if not XLSX_PATH.exists():
            pytest.skip(f"XLSX file not found: {XLSX_PATH}")

        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        ws = wb["Fig3"]

        for entry in golden_fixture["known_reuse"]:
            if not entry["full_row_match"]:
                continue

            r1 = entry["rows_a"][0]
            r2 = entry["rows_b"][0]

            # Read full row data (columns B-R = 2-18)
            actual_r1 = [ws.cell(r1, c).value for c in range(2, 19)]
            actual_r2 = [ws.cell(r2, c).value for c in range(2, 19)]

            assert actual_r1 == actual_r2, \
                f"Rows {r1} and {r2} marked as full_row_match but data differs"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not available")
    def test_consecutive_reuse_patterns(self, golden_fixture):
        """Verify consecutive CFU reuse patterns."""
        if not XLSX_PATH.exists():
            pytest.skip(f"XLSX file not found: {XLSX_PATH}")

        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        ws = wb["Fig3"]

        # Find the consecutive_cfu_reuse finding
        consecutive_finding = next(
            f for f in golden_fixture["expected_findings"]
            if f["type"] == "consecutive_cfu_reuse"
        )

        for pattern in consecutive_finding["patterns"]:
            rows_a = pattern["rows_a"]
            rows_b = pattern["rows_b"]

            # Verify CFU values match for each row pair
            for r1, r2 in zip(rows_a, rows_b):
                cfu_r1 = [ws.cell(r1, c).value for c in range(2, 6)]
                cfu_r2 = [ws.cell(r2, c).value for c in range(2, 6)]
                assert cfu_r1 == cfu_r2, \
                    f"Rows {r1} and {r2} should have identical CFU values"


class TestSemanticStructure:
    """Document the expected semantic structure."""

    def test_group_organization(self, golden_fixture):
        """Verify groups are organized by compound class and section."""
        groups = golden_fixture["groups"]

        # Group by compound class
        by_class = {}
        for g in groups:
            cc = g["compound_class"]
            if cc not in by_class:
                by_class[cc] = []
            by_class[cc].append(g)

        # Verify each compound class has groups
        for cc in ["Coumarins", "Benzenoids", "Flavanones", "Isoflavones", "Flavonol", "Chalcones"]:
            assert cc in by_class, f"Missing compound class: {cc}"
            assert len(by_class[cc]) > 0, f"Compound class {cc} has no groups"

        # Verify sections
        sections = {g["section"] for g in groups}
        assert len(sections) == 6, f"Expected 6 sections, got {len(sections)}"

    def test_cross_group_reuse_groups(self, golden_fixture):
        """Verify cross-group reuse involves distinct group pairs."""
        kr = golden_fixture["known_reuse"]

        # Extract group pairs
        group_pairs = [(e["group_a"], e["group_b"]) for e in kr]

        # Verify 3 main patterns
        # Pattern 1: Coumarins::2 <-> Flavanones::8 (3 rows)
        pattern1 = [(a, b) for a, b in group_pairs if a == "Coumarins::2" and b == "Flavanones::8"]
        assert len(pattern1) == 3, "Expected 3 rows shared between Coumarins::2 and Flavanones::8"

        # Pattern 2: Flavanones::7 <-> Isoflavones::11 (3 rows)
        pattern2 = [(a, b) for a, b in group_pairs if a == "Flavanones::7" and b == "Isoflavones::11"]
        assert len(pattern2) == 3, "Expected 3 rows shared between Flavanones::7 and Isoflavones::11"

        # Pattern 3: Other pairs (1 row each)
        other_patterns = [
            ("Benzenoids::wt", "Flavanones::WT"),
            ("Benzenoids::fob1", "Flavanones::fob1"),
            ("Flavanones::WT", "Flavonol::wt")
        ]
        for ga, gb in other_patterns:
            count = sum(1 for a, b in group_pairs if a == ga and b == gb)
            assert count == 1, f"Expected 1 row shared between {ga} and {gb}"
