"""Source Data LLM Verdict — Sheet-level batch false-positive adjudication.

Groups source_data findings by (workbook, sheet), reads XLSX column context,
and calls LLM (via AgentStepRunner / opencode) to judge each finding as
true_positive, false_positive, or uncertain.

Output: ``source_data/findings_verdict.json``

Design rationale
----------------
Deterministic pattern detectors (fixed_ratio, duplicate_row_vector, etc.)
flag many structurally normal data patterns as suspicious:

* Descriptive-statistics tables: Sum = Mean × N is a definitional relationship
* Index/time columns: parallel timelines have constant offsets
* scRNA-seq matrices: >60% zero-expression is biological dropout, not copy-paste
* Control groups: all-zero is a valid experimental outcome

Heuristic column-name enumeration cannot cover the open-ended vocabulary of
scientific data tables.  An LLM judge that sees the *actual column names*,
*sample values*, and *detected pattern* can understand the table's semantics
in one pass and return structured verdicts for every finding on the sheet.

Sheets are processed in parallel (ThreadPoolExecutor); each sheet is one LLM
call.  A typical paper produces 5-12 LLM calls, all parallel, so the wall-clock
latency equals a single LLM call.
"""

from __future__ import annotations

import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from engine.exceptions import AgentError
from engine.static_audit.paths import resolve_artifact_path
from engine.static_audit.tools.source_data_sheet_briefing import build_sheet_briefing

logger = logging.getLogger(__name__)

# ── Prompt template ──────────────────────────────────────────────────

_VERDICT_PROMPT = """\
You are a data-forensics investigator evaluating Source Data findings from a scientific paper.

The deterministic detector flagged statistical patterns (fixed ratio, fixed difference, \
duplicate columns, duplicate row vectors, paired ratio reuse, paired difference too narrow, \
row offset reuse) in the table below. Your task: decide whether EACH flagged pattern is a \
**false positive** (explainable by benign table structure), **true_positive** (genuine structural artifact), \
or **uncertain** (needs human review).

## Investigation Tools

You have access to the `source_data.query` tool for targeted investigation when you need to verify \
a hypothesis (e.g., cross-group duplicate patterns, specific column relationships). Use it when the \
available context is insufficient to make a confident verdict.

## Verdict Criteria

**false_positive** — when the relationship is *mechanically explainable* by any of:
- Definitional / arithmetic relationship: Sum = Mean × N, SD/SE derived from same data, Min/Mean/Max correlated by definition
- Unit conversion or normalization (e.g. fold-change, percentage, z-score derived from raw)
- Summary statistic pair: one column is a per-group aggregate of another
- Grouped experimental design: parallel timelines or parallel group columns that share a sampling frame
- Index/ID column: sequential row numbers, identifiers replicated across groups
- Zero-inflated technical artifact: scRNA-seq columns with >60% zero expression — zeros are biological dropout
- Constant column: a column with all identical values will have "fixed difference" with ANY other column

**uncertain** — when the pattern *might* be benign but you cannot rule out data manipulation:
- Duplicate measurements that could be independent replicates OR copy-paste
- Ratio patterns in measurement columns without clear definitional explanation
- Row vector matches where rows might represent independent samples

**true_positive** — ONLY for clear structural artifacts that ARE suspicious:
- Two measurement columns (not index/stats) with 100% identical values across all rows
- Ratio reuse in independently measured value columns (not descriptive stats)

## Priority Assignment

For each finding, assign a priority based on its impact on the paper's conclusions:

- **critical**: Directly supports the paper's main conclusion or key figure
- **high**: Affects a key figure or important secondary claim
- **medium**: Affects supplementary claims or supporting data
- **low**: Tangential to the paper's main narrative

Use the enriched_claims context (if provided) to understand which claims each finding relates to \
and how decisive those claims are for the paper's conclusions.

## Critical Rules
1. Default to **false_positive** when the relationship is mechanically explainable by table structure, column semantics, or standard scientific practice (normalization, unit conversion, summary statistics, grouped design). Only mark **uncertain** when the relationship lacks a clear mechanical explanation AND the columns represent independent experimental measurements
2. If the column labels clearly indicate a definitional relationship (Mean/Sum/Median of same data), mark false_positive
3. If columns represent independent experimental conditions or measurements, mark uncertain ONLY when the columns are confirmed to represent independent biological replicates measured under different conditions. If this cannot be confirmed, default to false_positive
4. Consider the WHOLE table structure — not just the flagged columns, but all columns together
5. Provide a clear, specific explanation for each verdict that cites actual data values from the table
6. Use source_data.query tool if you need to verify a hypothesis about cross-group patterns or column relationships

## Input
Sheet context is in the attached JSON file. It contains:
- `briefing`: compact sheet intelligence with:
  - `structure.group_count`: experimental groups detected (null if unknown)
  - `structure.total_data_rows`: total numeric data rows
  - `detected_patterns`: findings clustered by category, with count, max_risk, and \
`analysis_scope` (always "within-sheet" — detector only checked within this sheet)
  - `sample_data`: deduplicated raw data rows
- `columns`: column names and sample values from the actual spreadsheet
- `profile`: sheet-level statistics (cell counts, formula counts)
- `enriched_claims`: claims from the paper that reference this workbook/sheet (if available)

## Output
Return JSON matching this schema exactly:
{
  "sheet_verdict": "mostly_false_positive" | "mixed" | "mostly_uncertain",
  "sheet_pattern": "descriptive_statistics_table" | "index_time_columns" | \
"sparse_count_data" | "control_group_constants" | "measurement_data" | \
"long_format_data" | "mixed_structure" | "unknown",
  "explanation": "one-paragraph summary of why this sheet's patterns are/isn't suspicious",
  "findings": [
    {
      "id": "<finding_id from input>",
      "verdict": "true_positive" | "false_positive" | "uncertain",
      "confidence": 0.0 to 1.0,
      "priority": "critical" | "high" | "medium" | "low",
      "reason": "specific explanation citing data values",
      "benign_pattern": "<short label if false_positive, else null>" | null,
      "explanation": "one sentence"
    }
  ]
}

The `priority` and `reason` fields are optional but recommended. If omitted, priority defaults to "medium".

Mark findings as false_positive unless there is positive evidence that the relationship represents a genuine independent measurement anomaly.
"""


# ── XLSX context extraction ─────────────────────────────────────────


def _safe_value(cell_value: Any, max_len: int = 60) -> Any:
    """Convert a cell value to a JSON-safe primitive."""
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float, bool)):
        return cell_value
    s = str(cell_value)
    return s[:max_len] if len(s) > max_len else s


def read_xlsx_column_context(
    xlsx_path: Path,
    sheet_name: str,
    *,
    max_sample_rows: int = 5,
) -> dict[str, Any] | None:
    """Read column headers + sample values from an XLSX sheet.

    Handles multi-level merged-cell headers by:
    1. Scanning all rows until the first numeric-data row
    2. Collecting non-None header values per column (the "header hierarchy")
    3. Returning sample data rows after the header section

    Returns ``None`` if the file cannot be read (openpyxl missing, file
    corrupted, sheet not found).  The caller should treat ``None`` as
    "no XLSX context available" and still include the findings.
    """
    try:
        import openpyxl
    except ImportError:
        logger.debug("openpyxl not available — XLSX context unavailable")
        return None

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:  # Deliberately broad: openpyxl raises InvalidFileException, XML parsing errors, etc.
        logger.debug("Failed to open workbook for column context: %s", xlsx_path, exc_info=True)
        return None

    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]

        # Phase 1: scan for the header/data boundary.
        # A "data row" has at least one numeric (int/float) cell value.
        all_rows: list[list[Any]] = []
        first_data_row = -1
        for i, row in enumerate(ws.iter_rows(max_row=200, values_only=True)):
            row_vals = [_safe_value(c) for c in row]
            all_rows.append(row_vals)
            if first_data_row < 0 and any(
                isinstance(c, (int, float)) for c in row if c is not None
            ):
                first_data_row = i
            if first_data_row >= 0 and i >= first_data_row + max_sample_rows:
                break

        if not all_rows or first_data_row < 0:
            return None

        header_section = all_rows[:first_data_row]
        data_section = all_rows[first_data_row : first_data_row + max_sample_rows]

        # Drop all-None trailing header rows
        while header_section and all(v is None for v in header_section[-1]):
            header_section.pop()

        num_cols = max((len(r) for r in all_rows), default=0)

        # Phase 2: build per-column header hierarchy
        columns: list[dict[str, Any]] = []
        for col_idx in range(min(num_cols, 60)):
            hierarchy = []
            for row in header_section:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx])
                    if val and val not in hierarchy:
                        hierarchy.append(val)
            columns.append(
                {
                    "col_idx": col_idx,
                    "name": hierarchy[-1] if hierarchy else None,
                    "header_hierarchy": hierarchy,
                    "sample_values": [
                        row[col_idx] for row in data_section if col_idx < len(row)
                    ],
                }
            )

        return {
            "num_columns": num_cols,
            "num_rows_approx": ws.max_row or 0,
            "header_row_count": len(header_section),
            "data_start_row": first_data_row,
            "columns": columns,
        }
    except Exception:  # Deliberately broad: openpyxl cell access raises various undocumented exceptions
        logger.debug("Failed to read XLSX column context for %s/%s", xlsx_path.name, sheet_name, exc_info=True)
        return None
    finally:
        try:
            wb.close()
        except OSError:
            logger.debug("Failed to close workbook after reading column context: %s", xlsx_path, exc_info=True)


# ── Sheet context builder ────────────────────────────────────────────


def _build_sheet_context(
    workbook_name: str,
    sheet_name: str,
    findings: list[dict],
    source_data_dir: Path,
    profile: dict | None,
    enriched_claims: list[dict] | None = None,
) -> dict[str, Any]:
    """Assemble the per-sheet context payload sent to the LLM."""
    xlsx_path = source_data_dir / workbook_name

    xlsx_context = None
    if xlsx_path.exists():
        xlsx_context = read_xlsx_column_context(xlsx_path, sheet_name)

    profile_stats: dict[str, Any] = {}
    if profile:
        for wb in profile.get("workbooks", []):
            if wb.get("file_name") == workbook_name:
                for sh in wb.get("sheets", []):
                    if sh.get("name") == sheet_name:
                        profile_stats = {
                            k: sh[k]
                            for k in (
                                "cell_count",
                                "numeric_cell_count",
                                "formula_count",
                                "formula_sample",
                                "terminal_digit_counts",
                                "terminal_0_or_5_rate",
                                "duplicate_numeric_rows",
                            )
                            if k in sh
                        }
                        break
                break

    # ── Build sheet briefing (replaces per-finding context) ──
    briefing = build_sheet_briefing(
        workbook_name, sheet_name, findings, source_data_dir
    )

    return {
        "workbook": workbook_name,
        "sheet": sheet_name,
        "columns": (xlsx_context or {}).get("columns"),
        "header_row_count": (xlsx_context or {}).get("header_row_count"),
        "data_start_row": (xlsx_context or {}).get("data_start_row"),
        "xlsx_num_rows": (xlsx_context or {}).get("num_rows_approx"),
        "profile": profile_stats,
        "enriched_claims": enriched_claims or [],
        "briefing": briefing,
    }


# ── Schema validator ─────────────────────────────────────────────────


def _infer_sheet_verdict_from_findings(findings: list[dict[str, Any]]) -> str:
    verdicts = [str(finding.get("verdict") or "") for finding in findings]
    if any(verdict == "true_positive" for verdict in verdicts):
        return "mixed"
    if verdicts and all(verdict == "false_positive" for verdict in verdicts):
        return "mostly_false_positive"
    return "mostly_uncertain"


def _validate_verdict_output(data: Any) -> dict:
    """Validate LLM output against the verdict schema.

    Raises ``ValueError`` with a specific message on schema mismatch so
    AgentStepRunner can feed the error back to the LLM for retry.
    """
    if not isinstance(data, dict):
        raise ValueError("output is not a JSON object")
    if "findings" not in data:
        raise ValueError("missing 'findings'")
    if not isinstance(data["findings"], list):
        raise ValueError("'findings' is not a list")
    valid_priorities = ("critical", "high", "medium", "low")
    for fv in data["findings"]:
        if not isinstance(fv, dict):
            raise ValueError(f"finding entry is not an object: {fv!r}")
        if "id" not in fv:
            raise ValueError(f"finding entry missing 'id': {fv!r}")
        if fv.get("verdict") not in ("true_positive", "false_positive", "uncertain"):
            raise ValueError(
                f"invalid verdict {fv.get('verdict')!r} for {fv.get('id')}; "
                "must be true_positive | false_positive | uncertain"
            )
        # Validate optional priority field
        if "priority" in fv and fv["priority"] not in valid_priorities:
            raise ValueError(
                f"invalid priority {fv.get('priority')!r} for {fv.get('id')}; "
                "must be critical | high | medium | low"
            )
        # Validate optional reason field (must be string if present)
        if "reason" in fv and not isinstance(fv["reason"], str):
            raise ValueError(f"invalid reason type for {fv.get('id')}; must be string")

    if "sheet_verdict" not in data:
        data["sheet_verdict"] = _infer_sheet_verdict_from_findings(data["findings"])
        data["sheet_verdict_inferred"] = True
    elif data["sheet_verdict"] not in (
        "mostly_false_positive",
        "mixed",
        "mostly_uncertain",
    ):
        raise ValueError(
            f"invalid sheet_verdict: {data['sheet_verdict']!r}; "
            "must be mostly_false_positive | mixed | mostly_uncertain"
        )
    return data


# ── Single-sheet LLM call ────────────────────────────────────────────


def get_sheet_verdict(
    sheet_context: dict[str, Any],
    *,
    project_root: Path,
    env: dict[str, str],
    model: str = "dashscope/qwen3.7-plus",
    opencode_bin: str = "opencode",
    timeout_seconds: int = 300,
    max_retries: int = 2,
    log_dir: Path | None = None,
    workdir: Path | None = None,
) -> dict[str, Any]:
    """Call LLM to adjudicate all findings for a single sheet.

    Returns a verdict dict on success, or a fallback dict with
    ``verdict_status='failed'`` on failure.
    """
    from engine.investigation.agent_step_runner import AgentStepRunner

    workbook = sheet_context["workbook"]
    sheet = sheet_context["sheet"]
    tag = f"{workbook}__{sheet}".replace(" ", "_").replace("/", "_")

    runner = AgentStepRunner(
        project_root=project_root,
        model=model,
        opencode_bin=opencode_bin,
        env=env,
    )

    # Write context to a JSON file and attach via --file
    ctx_path = (log_dir or project_root / ".veritas-tmp") / f"verdict_ctx_{tag}.json"
    ctx_path.parent.mkdir(parents=True, exist_ok=True)
    ctx_path.write_text(json.dumps(sheet_context, indent=2, ensure_ascii=False))

    # Check if source_data.query tool is available
    has_query_tool = False
    if workdir:
        try:
            from engine.tools.registry import TOOLS

            has_query_tool = "source_data.query" in TOOLS
        except (ImportError, AttributeError):
            logger.debug("Failed to check tool registry for source_data.query", exc_info=True)
            has_query_tool = False

    # Build prompt with enriched context
    briefing = sheet_context.get("briefing", {})
    pattern_count = len(briefing.get("detected_patterns", []))
    finding_count = briefing.get("finding_count", 0)
    prompt_parts = [
        _VERDICT_PROMPT,
        "",
        "## This Sheet",
        f"Workbook: {workbook}",
        f"Sheet: {sheet}",
        f"Sheet briefing: {finding_count} findings clustered into {pattern_count} pattern categories",
    ]

    if has_query_tool:
        prompt_parts.append(
            "\nYou have access to the source_data.query tool for targeted investigation. "
            "Use it if you need to verify cross-group patterns or specific column relationships."
        )

    enriched_claims = sheet_context.get("enriched_claims", [])
    if enriched_claims:
        prompt_parts.append(
            f"\nEnriched claims context: {len(enriched_claims)} claims reference this sheet. "
            "See the attached JSON for details including claim_decisiveness and expected_source_data."
        )

    prompt_parts.append(
        "\nRead the attached JSON file for full table structure and findings. "
        f"Context path: {ctx_path}. "
        "Return your verdict JSON."
    )

    prompt = "\n".join(prompt_parts)

    try:
        result = runner.run(
            role=f"verdict_{tag}",
            prompt=prompt,
            output_validator=_validate_verdict_output,
            timeout_seconds=timeout_seconds,
            max_retries=max_retries,
            context_pack_path=ctx_path,
            log_dir=log_dir,
        )
    finally:
        try:
            ctx_path.unlink(missing_ok=True)
        except OSError:
            logger.debug("Failed to clean up verdict context file: %s", ctx_path, exc_info=True)

    if result.status == "success" and result.output:
        output = dict(result.output)
        output["verdict_status"] = "success"
        output["runtime_seconds"] = result.runtime_seconds
        return output

    # Failure fallback — keep findings unjudged
    detail = result.metadata.get("last_detail", "") if result.metadata else ""
    logger.warning("LLM verdict failed for %s/%s: %s", workbook, sheet, detail)
    return {
        "sheet_verdict": "unknown",
        "sheet_pattern": "unknown",
        "verdict_status": "failed",
        "explanation": f"LLM verdict call failed: {detail}",
        "findings": [
            {
                "id": p.get("category"),
                "verdict": "uncertain",
                "confidence": 0.0,
                "priority": "medium",
                "explanation": "LLM verdict unavailable (default fallback — not LLM-judged)",
            }
            for p in (sheet_context.get("briefing") or {}).get("detected_patterns", [])
        ],
    }


# ── Cluster → finding verdict expansion ──────────────────────────────


def _expand_cluster_verdicts(
    cluster_verdicts: list[dict[str, Any]],
    briefing: dict[str, Any],
) -> list[dict[str, Any]]:
    """Expand cluster-level verdicts back to per-finding verdicts.

    The LLM judges pattern clusters (detected_patterns in the briefing).
    This function maps each cluster verdict back to all individual findings
    in that cluster, so downstream consumers get per-finding verdicts.
    """
    patterns = briefing.get("detected_patterns") or []
    finding_count = briefing.get("finding_count", 0)

    # Build cluster_id → cluster verdict mapping
    verdict_by_category: dict[str, dict[str, Any]] = {}
    for cv in cluster_verdicts:
        cat_id = cv.get("id", "")
        verdict_by_category[cat_id] = cv

    # If the LLM returned individual finding verdicts (not cluster-level),
    # pass them through unchanged
    if len(cluster_verdicts) == finding_count:
        return cluster_verdicts

    # If no briefing patterns, return cluster verdicts as-is
    if not patterns:
        return cluster_verdicts

    # Expand: each finding in a cluster gets the cluster's verdict
    expanded: list[dict[str, Any]] = []
    for pattern in patterns:
        cat = pattern.get("category", "unknown")
        count = pattern.get("count", 1)
        cv = verdict_by_category.get(cat, {})
        for i in range(count):
            expanded.append(
                {
                    "id": f"{cat}-{i + 1:04d}",
                    "category": cat,
                    "verdict": cv.get("verdict", "uncertain"),
                    "confidence": cv.get("confidence", 0.0),
                    "priority": cv.get("priority", "medium"),
                    "explanation": cv.get("explanation", ""),
                    "benign_pattern": cv.get("benign_pattern"),
                    "from_cluster": True,
                }
            )
    return expanded


# ── Grouping logic ───────────────────────────────────────────────────


def _group_findings_by_sheet(
    findings_data: dict | None,
    pair_forensics_data: dict | None,
) -> dict[tuple[str, str], list[dict]]:
    """Merge findings from both sources, grouped by (workbook, sheet)."""
    grouped: dict[tuple[str, str], list[dict]] = {}

    def _add(finding: dict) -> None:
        wb = finding.get("workbook", "")
        sh = finding.get("sheet", "")
        if not wb or not sh:
            return
        grouped.setdefault((wb, sh), []).append(finding)

    for f in (findings_data or {}).get("findings", []):
        _add(f)
    for f in (pair_forensics_data or {}).get("findings", []):
        _add(f)

    return grouped


# ── Enriched claims filtering ────────────────────────────────────────


def _filter_claims_for_sheet(
    claims: list[dict],
    workbook_name: str,
    sheet_name: str,
) -> list[dict]:
    """Filter enriched claims to those that reference the given workbook/sheet.

    A claim references a sheet if its `expected_source_data` field contains
    the workbook name (case-insensitive) or sheet name.

    Returns a list of compact claim summaries suitable for LLM context.
    """
    if not claims:
        return []

    wb_lower = workbook_name.lower()
    sh_lower = sheet_name.lower()

    matching: list[dict] = []
    for claim in claims:
        expected_sd = claim.get("expected_source_data", [])
        if not expected_sd:
            continue
        # Check if any expected_source_data entry references this workbook or sheet
        for ref in expected_sd:
            ref_lower = str(ref).lower()
            if wb_lower in ref_lower or sh_lower in ref_lower:
                # Extract compact claim info for LLM context
                matching.append(
                    {
                        "claim_id": claim.get("claim_id"),
                        "claim_text": claim.get("claim_text", "")[:200],
                        "claim_type": claim.get("claim_type"),
                        "claim_decisiveness": claim.get("claim_decisiveness", "medium"),
                        "figure_refs": claim.get("figure_refs", []),
                        "expected_source_data": expected_sd,
                    }
                )
                break  # Only add claim once even if multiple refs match

    return matching


# ── Main entry point ─────────────────────────────────────────────────


def run_source_data_verdict(
    workdir: Path,
    *,
    source_data_dir: Path,
    project_root: Path,
    env: dict[str, str],
    model: str = "dashscope/qwen3.7-plus",
    opencode_bin: str = "opencode",
    force: bool = False,
    progress: Any = None,
    timeout_seconds: int = 300,
    max_retries: int = 2,
    max_workers: int = 4,
) -> dict[str, Any]:
    """Run LLM verdict on all source-data findings.

    Reads ``source_data/findings.json`` and ``source_data/pair_forensics.json``,
    groups by sheet, calls LLM in parallel for each sheet, and writes
    ``source_data/findings_verdict.json``.

    Returns the full verdict result dict.
    """
    output_path = resolve_artifact_path(workdir, "source_data_findings_verdict.json")

    if output_path.exists() and not force:
        logger.info("Reusing existing findings_verdict.json")
        return json.loads(output_path.read_text(encoding="utf-8"))

    findings_path = resolve_artifact_path(workdir, "source_data_findings.json")
    pair_path = resolve_artifact_path(workdir, "source_data_pair_forensics.json")

    findings_data = (
        json.loads(findings_path.read_text(encoding="utf-8"))
        if findings_path.exists()
        else None
    )
    pair_forensics_data = (
        json.loads(pair_path.read_text(encoding="utf-8"))
        if pair_path.exists()
        else None
    )

    if not findings_data and not pair_forensics_data:
        result: dict[str, Any] = {
            "schema_version": "1.0",
            "verdict_status": "skipped",
            "explanation": "No source data findings to adjudicate.",
            "sheets": [],
            "summary": {
                "total_sheets": 0,
                "total_findings": 0,
                "true_positive": 0,
                "false_positive": 0,
                "uncertain": 0,
            },
        }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(result, indent=2, ensure_ascii=False))
        return result

    grouped = _group_findings_by_sheet(findings_data, pair_forensics_data)

    # Profile for per-sheet statistics
    profile_path = resolve_artifact_path(workdir, "source_data_profile.json")
    profile = (
        json.loads(profile_path.read_text(encoding="utf-8"))
        if profile_path.exists()
        else None
    )

    log_dir = resolve_artifact_path(workdir, "logs")
    log_dir.mkdir(parents=True, exist_ok=True)

    # Load enriched claims from agent_claim_extractor artifact
    enriched_claims: list[dict] = []
    claims_path = resolve_artifact_path(workdir, "agent_claim_extractor.json")
    if claims_path.exists():
        try:
            claims_data = json.loads(claims_path.read_text(encoding="utf-8"))
            enriched_claims = claims_data.get("claims", [])
            logger.info(
                "Loaded %d enriched claims for verdict context", len(enriched_claims)
            )
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            logger.warning("Failed to load enriched claims: %s", e)

    # Build sheet contexts
    sheet_contexts: list[dict[str, Any]] = []
    for (wb, sh), fs in sorted(grouped.items()):
        # Filter enriched claims that reference this workbook/sheet
        sheet_claims = _filter_claims_for_sheet(enriched_claims, wb, sh)
        ctx = _build_sheet_context(wb, sh, fs, source_data_dir, profile, sheet_claims)
        sheet_contexts.append(ctx)

    # Parallel LLM calls
    sheet_verdicts: list[dict[str, Any]] = []
    workers = min(max_workers, len(sheet_contexts)) or 1
    logger.info(
        "Running LLM verdict on %d sheets (%d workers)",
        len(sheet_contexts),
        workers,
    )

    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_to_ctx = {
            pool.submit(
                get_sheet_verdict,
                ctx,
                project_root=project_root,
                env=env,
                model=model,
                opencode_bin=opencode_bin,
                timeout_seconds=timeout_seconds,
                max_retries=max_retries,
                log_dir=log_dir,
                workdir=workdir,
            ): ctx
            for ctx in sheet_contexts
        }
        for future in as_completed(future_to_ctx):
            ctx = future_to_ctx[future]
            try:
                verdict = future.result()
            except (OSError, AgentError) as exc:
                logger.warning(
                    "Verdict exception for %s/%s: %s",
                    ctx["workbook"],
                    ctx["sheet"],
                    exc,
                )
                verdict = {
                    "sheet_verdict": "unknown",
                    "sheet_pattern": "unknown",
                    "verdict_status": "failed",
                    "explanation": f"Exception: {exc}",
                    "findings": [
                        {
                            "id": p.get("category"),
                            "verdict": "uncertain",
                            "confidence": 0.0,
                            "priority": "medium",
                            "explanation": "LLM verdict unavailable (default fallback — not LLM-judged)",
                        }
                        for p in (ctx.get("briefing") or {}).get(
                            "detected_patterns", []
                        )
                    ],
                }
            # Ensure workbook/sheet are in the verdict
            verdict["workbook"] = ctx["workbook"]
            verdict["sheet"] = ctx["sheet"]
            verdict["finding_count"] = (ctx.get("briefing") or {}).get(
                "finding_count", 0
            )

            # Expand cluster-level verdicts back to per-finding verdicts
            # The LLM judges pattern clusters; we map back to individual findings
            verdict["findings"] = _expand_cluster_verdicts(
                verdict.get("findings", []),
                ctx.get("briefing", {}),
            )
            sheet_verdicts.append(verdict)

    # Stable ordering by (workbook, sheet)
    sheet_verdicts.sort(key=lambda v: (v.get("workbook", ""), v.get("sheet", "")))

    # Summary
    tp = fp = un = 0
    for sv in sheet_verdicts:
        for fv in sv.get("findings", []):
            v = fv.get("verdict", "uncertain")
            if v == "true_positive":
                tp += 1
            elif v == "false_positive":
                fp += 1
            else:
                un += 1

    result = {
        "schema_version": "1.0",
        "created_by": "engine/static_audit/tools/source_data_verdict.py",
        "model": model,
        "sheets": sheet_verdicts,
        "summary": {
            "total_sheets": len(sheet_verdicts),
            "total_findings": tp + fp + un,
            "true_positive": tp,
            "false_positive": fp,
            "uncertain": un,
            "failed_sheets": sum(
                1 for sv in sheet_verdicts if sv.get("verdict_status") == "failed"
            ),
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, indent=2, ensure_ascii=False))
    logger.info(
        "Verdict complete: %d sheets, %d TP, %d FP, %d uncertain",
        len(sheet_verdicts),
        tp,
        fp,
        un,
    )
    return result
