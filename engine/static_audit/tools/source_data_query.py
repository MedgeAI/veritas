"""Semantic Source Data query tool.

Provides high-level deterministic queries over Source Data workbooks:
- compare_groups: compare data between two groups in a sheet
- extract_block: extract a data block for a specific group
- find_cross_group_reuse: find rows shared across groups
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]

try:
    from engine.static_audit.paths import resolve_artifact_path
except ImportError:
    resolve_artifact_path = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_row_range(spec: str) -> tuple[int, int] | None:
    """Parse a row range spec like '2-15' into (start, end) 1-based inclusive."""
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", spec.strip())
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def _parse_column_range(spec: str) -> tuple[int, int] | None:
    """Parse a column range spec like 'B-E' or '2-5' into (start, end) 1-based.

    Supports Excel-style letters (A=1, B=2, ...) or numeric indices.
    """
    spec = spec.strip()
    # Try letter-based: "B-E"
    m = re.fullmatch(r"([A-Za-z]+)\s*-\s*([A-Za-z]+)", spec)
    if m:
        return _col_letter_to_index(m.group(1)), _col_letter_to_index(m.group(2))
    # Try numeric: "2-5"
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", spec)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def _col_letter_to_index(letters: str) -> int:
    """Convert column letters (A, B, ..., Z, AA, ...) to 1-based index."""
    result = 0
    for ch in letters.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _index_to_col_letter(index: int) -> str:
    """Convert 1-based column index to letter."""
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _read_cell_numeric(ws: Any, row: int, col: int) -> float | None:
    """Read a cell value as float if numeric, else None."""
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val)
        except ValueError:
            return None
    return None


def _find_group_rows_by_label(
    ws: Any, label: str, label_col: int = 1
) -> tuple[int, int] | None:
    """Find rows belonging to a text-based group label.

    Scans label_col for the group label, then returns the contiguous
    block of non-empty data rows below it until the next label or blank row.
    """
    max_row = ws.max_row
    # Find header/label row
    start_row = None
    for r in range(1, max_row + 1):
        cell_val = ws.cell(row=r, column=label_col).value
        if cell_val is not None and str(cell_val).strip().lower() == label.lower():
            start_row = r + 1
            break
    if start_row is None:
        return None
    # Find end: next label or 3 consecutive empty rows
    end_row = start_row
    empty_count = 0
    for r in range(start_row, max_row + 1):
        cell_val = ws.cell(row=r, column=label_col).value
        if cell_val is None or str(cell_val).strip() == "":
            empty_count += 1
            if empty_count >= 3:
                break
        else:
            empty_count = 0
            end_row = r
    if end_row < start_row:
        return None
    return start_row, end_row


def _resolve_group(ws: Any, group_spec: str) -> tuple[int, int]:
    """Resolve a group spec to (start_row, end_row). Supports range or label."""
    rng = _parse_row_range(group_spec)
    if rng is not None:
        return rng
    result = _find_group_rows_by_label(ws, group_spec)
    if result is not None:
        return result
    raise ValueError(f"Cannot resolve group: {group_spec!r}")


def _get_column_values(
    ws: Any, start_row: int, end_row: int, col: int
) -> list[float | None]:
    """Read numeric values for a column across a row range."""
    return [_read_cell_numeric(ws, r, col) for r in range(start_row, end_row + 1)]


def _compare_two_columns(
    values_a: list[float | None], values_b: list[float | None]
) -> dict[str, Any]:
    """Compare two numeric column value lists, detecting patterns."""
    exact = 0
    fixed_diffs: list[float] = []
    fixed_ratios: list[float] = []
    total_pairs = 0

    for va, vb in zip(values_a, values_b):
        if va is None or vb is None:
            continue
        total_pairs += 1
        if va == vb:
            exact += 1
        diff = vb - va
        fixed_diffs.append(diff)
        if va != 0:
            fixed_ratios.append(vb / va)

    result: dict[str, Any] = {"total_pairs": total_pairs, "exact_matches": exact}

    if total_pairs == 0:
        result["pattern"] = "no_data"
        return result

    # Detect fixed difference
    if fixed_diffs:
        unique_diffs = set(round(d, 10) for d in fixed_diffs)
        if len(unique_diffs) == 1 and total_pairs > 1:
            result["pattern"] = "fixed_difference"
            result["fixed_difference"] = fixed_diffs[0]
            result["support"] = 1.0
            return result

    # Detect fixed ratio
    if fixed_ratios:
        unique_ratios = set(round(r, 10) for r in fixed_ratios)
        if len(unique_ratios) == 1 and total_pairs > 1:
            result["pattern"] = "fixed_ratio"
            result["fixed_ratio"] = fixed_ratios[0]
            result["support"] = 1.0
            return result

    # Near-fixed difference (>90% same diff)
    if len(fixed_diffs) > 2:
        diff_counts: dict[float, int] = {}
        for d in fixed_diffs:
            rd = round(d, 6)
            diff_counts[rd] = diff_counts.get(rd, 0) + 1
        most_common_diff = max(diff_counts, key=diff_counts.get)  # type: ignore[arg-type]
        count = diff_counts[most_common_diff]
        rate = count / len(fixed_diffs)
        if rate >= 0.9 and count > 2:
            result["pattern"] = "near_fixed_difference"
            result["fixed_difference"] = most_common_diff
            result["support"] = round(rate, 4)
            return result

    result["pattern"] = "none"
    return result


# ---------------------------------------------------------------------------
# Query implementations
# ---------------------------------------------------------------------------


def _query_compare_groups(
    ws: Any,
    group_a_spec: str,
    group_b_spec: str,
    column_block_spec: str,
) -> dict[str, Any]:
    """Compare data between two groups in a sheet."""
    row_a_start, row_a_end = _resolve_group(ws, group_a_spec)
    row_b_start, row_b_end = _resolve_group(ws, group_b_spec)
    col_start, col_end = _parse_column_range(column_block_spec)  # type: ignore[misc]
    if col_start is None:
        raise ValueError(f"Invalid column_block: {column_block_spec!r}")

    per_column: list[dict[str, Any]] = []
    for col in range(col_start, col_end + 1):
        col_letter = _index_to_col_letter(col)
        header = ws.cell(row=1, column=col).value
        values_a = _get_column_values(ws, row_a_start, row_a_end, col)
        values_b = _get_column_values(ws, row_b_start, row_b_end, col)
        comparison = _compare_two_columns(values_a, values_b)
        comparison["column"] = col_letter
        comparison["column_header"] = str(header) if header else None
        per_column.append(comparison)

    return {
        "query_type": "compare_groups",
        "group_a": {"spec": group_a_spec, "rows": [row_a_start, row_a_end]},
        "group_b": {"spec": group_b_spec, "rows": [row_b_start, row_b_end]},
        "column_block": {
            "spec": column_block_spec,
            "range": [col_start, col_end],
        },
        "columns": per_column,
        "summary": {
            "total_columns": len(per_column),
            "exact_match_columns": sum(
                1
                for c in per_column
                if c.get("pattern") == "none" and c.get("exact_matches", 0) > 0
            ),
            "fixed_difference_columns": sum(
                1 for c in per_column if c.get("pattern") == "fixed_difference"
            ),
            "fixed_ratio_columns": sum(
                1 for c in per_column if c.get("pattern") == "fixed_ratio"
            ),
        },
    }


def _query_extract_block(
    ws: Any,
    group_spec: str,
    columns_spec: str,
) -> dict[str, Any]:
    """Extract a data block for a specific group."""
    row_start, row_end = _resolve_group(ws, group_spec)
    col_start, col_end = _parse_column_range(columns_spec)  # type: ignore[misc]
    if col_start is None:
        raise ValueError(f"Invalid columns: {columns_spec!r}")

    headers = []
    for col in range(col_start, col_end + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(str(h) if h else _index_to_col_letter(col))

    rows_data: list[list[Any]] = []
    for r in range(row_start, row_end + 1):
        row_vals: list[Any] = []
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=col)
            val = cell.value
            row_vals.append(val)
        rows_data.append(row_vals)

    return {
        "query_type": "extract_block",
        "group": {"spec": group_spec, "rows": [row_start, row_end]},
        "columns": {
            "spec": columns_spec,
            "range": [col_start, col_end],
            "headers": headers,
        },
        "data": rows_data,
        "row_count": len(rows_data),
        "column_count": len(headers),
    }


def _query_find_cross_group_reuse(
    ws: Any,
    column_block_spec: str,
    groups: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    """Find rows that appear in multiple groups.

    groups: optional list of {"name": ..., "rows": "start-end"} dicts.
    If not provided, scans for text-separated groups automatically.
    """
    col_start, col_end = _parse_column_range(column_block_spec)  # type: ignore[misc]
    if col_start is None:
        raise ValueError(f"Invalid column_block: {column_block_spec!r}")

    # Resolve groups
    resolved_groups: list[tuple[str, int, int]] = []
    if groups:
        for g in groups:
            name = g.get("name", "")
            row_spec = g.get("rows", "")
            rng = _parse_row_range(row_spec)
            if rng is None:
                result = _find_group_rows_by_label(ws, row_spec)
                if result is None:
                    continue
                rng = result
            resolved_groups.append((name, rng[0], rng[1]))
    else:
        # Auto-detect groups: scan column 1 for text labels
        label_col = 1
        max_row = ws.max_row
        current_label = None
        current_start = None
        for r in range(2, max_row + 1):  # skip header row
            cell_val = ws.cell(row=r, column=label_col).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.strip():
                text = cell_val.strip()
                # Heuristic: short text that looks like a label (no digits as first char)
                if not text[0].isdigit() and len(text) < 50:
                    if current_label is not None and current_start is not None:
                        resolved_groups.append(
                            (current_label, current_start, r - 1)
                        )
                    current_label = text
                    current_start = r + 1
        if current_label is not None and current_start is not None:
            resolved_groups.append((current_label, current_start, max_row))

    if len(resolved_groups) < 2:
        return {
            "query_type": "find_cross_group_reuse",
            "column_block": {"spec": column_block_spec, "range": [col_start, col_end]},
            "groups_found": len(resolved_groups),
            "shared_rows": [],
            "summary": {"total_shared": 0},
        }

    # Build row fingerprints per group
    shared_rows: list[dict[str, Any]] = []
    # Compare all pairs of groups
    for i in range(len(resolved_groups)):
        for j in range(i + 1, len(resolved_groups)):
            name_i, start_i, end_i = resolved_groups[i]
            name_j, start_j, end_j = resolved_groups[j]
            # Build fingerprints
            fps_i: dict[str, int] = {}
            for r in range(start_i, end_i + 1):
                fp = _row_fingerprint(ws, r, col_start, col_end)
                if fp:
                    fps_i[fp] = r
            for r in range(start_j, end_j + 1):
                fp = _row_fingerprint(ws, r, col_start, col_end)
                if fp and fp in fps_i:
                    # Collect shared values
                    values = []
                    for col in range(col_start, col_end + 1):
                        val = _read_cell_numeric(ws, fps_i[fp], col)
                        values.append(val)
                    shared_rows.append(
                        {
                            "row_a": fps_i[fp],
                            "row_b": r,
                            "group_a": name_i,
                            "group_b": name_j,
                            "shared_values": values,
                        }
                    )

    return {
        "query_type": "find_cross_group_reuse",
        "column_block": {"spec": column_block_spec, "range": [col_start, col_end]},
        "groups_found": len(resolved_groups),
        "shared_rows": shared_rows,
        "summary": {"total_shared": len(shared_rows)},
    }


def _row_fingerprint(
    ws: Any, row: int, col_start: int, col_end: int
) -> str | None:
    """Create a fingerprint string for a row's numeric values."""
    vals: list[str] = []
    has_numeric = False
    for col in range(col_start, col_end + 1):
        v = _read_cell_numeric(ws, row, col)
        if v is not None:
            has_numeric = True
            vals.append(f"{v:.10g}")
        else:
            vals.append("_")
    if not has_numeric:
        return None
    return "|".join(vals)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def run_source_data_query(
    workdir: Path,
    *,
    source_data_dir: Path,
    query_type: str,
    workbook: str,
    sheet: str,
    group_a: str = "",
    group_b: str = "",
    column_block: str = "",
    columns: str = "",
    group: str = "",
    groups: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    """Run a semantic Source Data query and write JSON artifact.

    Parameters
    ----------
    workdir : Path
        Audit working directory for output.
    source_data_dir : Path
        Directory containing Source Data XLSX files.
    query_type : str
        One of: compare_groups, extract_block, find_cross_group_reuse.
    workbook : str
        Workbook filename (relative to source_data_dir).
    sheet : str
        Sheet name.
    group_a, group_b : str
        Row ranges (e.g. "2-15") or text labels for compare_groups.
    column_block : str
        Column range (e.g. "B-E" or "2-5").
    columns : str
        Column range for extract_block.
    group : str
        Group spec for extract_block.
    groups : list[dict] | None
        Group definitions for find_cross_group_reuse.
    """
    if load_workbook is None:
        return {
            "query_type": query_type,
            "status": "error",
            "error": "openpyxl is not installed",
        }

    wb_path = Path(source_data_dir) / workbook
    if not wb_path.exists():
        return {
            "query_type": query_type,
            "status": "error",
            "error": f"Workbook not found: {wb_path}",
        }

    wb = load_workbook(str(wb_path), read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return {
                "query_type": query_type,
                "status": "error",
                "error": f"Sheet {sheet!r} not found. Available: {wb.sheetnames}",
            }
        ws = wb[sheet]

        if query_type == "compare_groups":
            result = _query_compare_groups(ws, group_a, group_b, column_block)
        elif query_type == "extract_block":
            result = _query_extract_block(ws, group or group_a, columns or column_block)
        elif query_type == "find_cross_group_reuse":
            result = _query_find_cross_group_reuse(ws, column_block, groups)
        else:
            return {
                "query_type": query_type,
                "status": "error",
                "error": f"Unknown query_type: {query_type!r}",
            }
    finally:
        wb.close()

    result["workbook"] = workbook
    result["sheet"] = sheet
    result["status"] = "ok"

    # Write artifact
    if resolve_artifact_path is not None:
        output_path = resolve_artifact_path(workdir, "source_data_query.json")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(result, indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
        logger.info("Source data query result written to %s", output_path)

    return result
