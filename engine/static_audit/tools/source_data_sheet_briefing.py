"""Sheet Briefing — compact structural intelligence for Agent context.

One function replaces three separate context-building patterns:
1. Sheet structure analysis (groups, columns, data layout)
2. Finding clustering (category → count + risk + representative rows)
3. Sample data aggregation (deduplicated across findings, capped)

Output: one JSON briefing per sheet, consumed by verdict LLM and
investigation planner as the single source of sheet-level information.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]


# ── Risk ordering ──────────────────────────────────────────────────────

_RISK_RANK = {"critical": 4, "high": 3, "medium": 2, "low": 1, "unknown": 0}


def _max_risk(risks: list[str]) -> str:
    if not risks:
        return "unknown"
    return max(risks, key=lambda r: _RISK_RANK.get(r, 0))


# ── Sheet structure analysis ──────────────────────────────────────────


def _detect_groups(ws: Any, max_row: int) -> dict[str, Any]:
    """Detect experimental groups by scanning column A for text separators.

    Returns {group_count, group_labels, total_data_rows} or
    {group_count: null, ...} if no groups detected.
    """
    labels: list[str] = []
    data_rows = 0
    scan_limit = min(max_row, 2000)

    for row in ws.iter_rows(min_row=2, max_row=scan_limit, min_col=1, max_col=1, values_only=True):
        val = row[0] if row else None
        if val is None:
            continue
        if isinstance(val, str) and val.strip():
            text = val.strip()
            if not text[0].isdigit() and len(text) < 80:
                labels.append(text)
        elif isinstance(val, (int, float)):
            data_rows += 1

    if not labels:
        return {
            "group_count": None,
            "group_labels": [],
            "total_data_rows": data_rows,
        }

    # Deduplicate labels while preserving order
    seen: set[str] = set()
    unique_labels: list[str] = []
    for label in labels:
        if label not in seen:
            seen.add(label)
            unique_labels.append(label)

    return {
        "group_count": len(unique_labels),
        "group_labels": unique_labels[:50],  # cap labels
        "total_data_rows": data_rows,
    }


def _detect_column_blocks(ws: Any, max_col: int) -> list[dict[str, str]]:
    """Detect column blocks from header row."""
    scan_cols = min(max_col, 60)
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=scan_cols, values_only=True):
        for val in row:
            headers.append(str(val).strip() if val is not None else "")
        break

    if not headers:
        return []

    # Simple block detection: group consecutive non-empty headers
    blocks: list[dict[str, str]] = []
    block_start = -1
    for i, h in enumerate(headers):
        if h and block_start < 0:
            block_start = i
        elif not h and block_start >= 0:
            block_end = i - 1
            _add_block(blocks, block_start, block_end, headers)
            block_start = -1
    # Close last block
    if block_start >= 0:
        _add_block(blocks, block_start, len(headers) - 1, headers)

    return blocks[:20]  # cap blocks


def _add_block(
    blocks: list[dict[str, str]],
    start: int,
    end: int,
    headers: list[str],
) -> None:
    start_letter = _col_index_to_letter(start + 1)
    end_letter = _col_index_to_letter(end + 1)
    width = end - start + 1
    block_headers = [h for h in headers[start : end + 1] if h]
    hint = f"{width}-col block"
    if block_headers:
        hint = f"{', '.join(block_headers[:3])}"
        if len(block_headers) > 3:
            hint += f" (+{len(block_headers) - 3} more)"
    blocks.append({"columns": f"{start_letter}-{end_letter}", "hint": hint})


def _col_index_to_letter(index: int) -> str:
    """Convert 1-based column index to Excel letter."""
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _analyze_sheet_structure(
    xlsx_path: Path, sheet_name: str
) -> dict[str, Any] | None:
    """Read XLSX and return structural analysis for one sheet."""
    if load_workbook is None:
        return None
    if not xlsx_path.exists():
        return None

    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        logger.debug("Failed to open workbook for structure analysis: %s", xlsx_path, exc_info=True)
        return None

    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        groups = _detect_groups(ws, max_row)
        column_blocks = _detect_column_blocks(ws, max_col)

        return {
            "group_count": groups["group_count"],
            "group_labels": groups["group_labels"][:20],
            "total_data_rows": groups["total_data_rows"],
            "column_count": max_col,
            "column_blocks": column_blocks,
        }
    except Exception:
        logger.debug("Failed to analyze sheet structure for %s/%s", xlsx_path.name, sheet_name, exc_info=True)
        return None
    finally:
        try:
            wb.close()
        except Exception:
            logger.debug("Failed to close workbook after analysis: %s", xlsx_path, exc_info=True)


# ── Finding clustering ────────────────────────────────────────────────


def _cluster_findings(findings: list[dict]) -> list[dict[str, Any]]:
    """Cluster findings by category, aggregating count/risk/representative rows."""
    by_category: dict[str, list[dict]] = defaultdict(list)
    for f in findings:
        cat = f.get("category", "unknown")
        by_category[cat].append(f)

    clusters: list[dict[str, Any]] = []
    for cat, items in sorted(by_category.items(), key=lambda x: -len(x[1])):
        risks = [str(f.get("risk_level", "unknown")) for f in items]

        # Collect representative row pairs from various finding formats
        row_pairs: list[list[int]] = []
        for f in items[:20]:  # scan up to 20 for representatives
            pairs = _extract_row_pairs(f)
            row_pairs.extend(pairs)

        cluster: dict[str, Any] = {
            "category": cat,
            "count": len(items),
            "max_risk": _max_risk(risks),
            "analysis_scope": "within-sheet",
        }
        if row_pairs:
            cluster["representative_row_pairs"] = row_pairs[:5]
        clusters.append(cluster)

    return clusters


def _extract_row_pairs(finding: dict) -> list[list[int]]:
    """Extract representative row pairs from a finding dict."""
    pairs: list[list[int]] = []

    # duplicate_row_vector: sample_rows or duplicate_rows
    for key in ("sample_rows", "duplicate_rows"):
        rows = finding.get(key)
        if isinstance(rows, list) and len(rows) >= 2:
            pairs.append([int(rows[0]), int(rows[1])])
            break

    # paired_ratio_reuse / row_offset: matched_pairs or support_rows
    matched = finding.get("matched_pairs")
    if isinstance(matched, list) and matched:
        for mp in matched[:3]:
            if isinstance(mp, dict):
                r1 = mp.get("left_row") or mp.get("row_a")
                r2 = mp.get("right_row") or mp.get("row_b")
                if r1 is not None and r2 is not None:
                    pairs.append([int(r1), int(r2)])

    # cross_block: block_a_rows + block_b_rows
    block_a = finding.get("block_a_rows")
    block_b = finding.get("block_b_rows")
    if isinstance(block_a, list) and isinstance(block_b, list) and block_a and block_b:
        pairs.append([int(block_a[0]), int(block_b[0])])

    # sample_pairs (fixed_difference/ratio)
    sp = finding.get("sample_pairs")
    if isinstance(sp, list) and sp:
        for item in sp[:3]:
            if isinstance(item, dict):
                r = item.get("row")
                if r is not None:
                    pairs.append([int(r)])

    return pairs


# ── Sample data aggregation ───────────────────────────────────────────


def _aggregate_sample_data(
    findings: list[dict], max_rows: int = 50
) -> tuple[list[dict[str, Any]], str]:
    """Deduplicate raw_data_samples across findings, cap at max_rows."""
    rows_by_number: dict[int, dict[str, Any]] = {}
    for f in findings:
        for sample in f.get("raw_data_samples") or []:
            row_num = sample.get("row")
            if row_num is not None and row_num not in rows_by_number:
                rows_by_number[row_num] = sample

    total = len(rows_by_number)
    capped = [rows_by_number[r] for r in sorted(rows_by_number)[:max_rows]]
    note = (
        f"{total} unique rows across all findings"
        + (f" (capped at {max_rows})" if total > max_rows else "")
    )
    return capped, note


# ── Main entry points ─────────────────────────────────────────────────


def build_sheet_briefing(
    workbook_name: str,
    sheet_name: str,
    findings: list[dict],
    source_data_dir: Path,
) -> dict[str, Any]:
    """Build a compact sheet briefing for Agent consumption.

    Combines structure analysis, finding clustering, and sample data
    into one high-density summary.
    """
    xlsx_path = source_data_dir / workbook_name
    structure = _analyze_sheet_structure(xlsx_path, sheet_name)
    detected_patterns = _cluster_findings(findings)
    sample_data, sample_note = _aggregate_sample_data(findings)

    briefing: dict[str, Any] = {
        "sheet": sheet_name,
        "workbook": workbook_name,
        "finding_count": len(findings),
    }
    if structure:
        briefing["structure"] = structure
    else:
        briefing["structure"] = {"group_count": None, "total_data_rows": None}
    briefing["detected_patterns"] = detected_patterns
    if sample_data:
        briefing["sample_data"] = sample_data
        briefing["sample_data_note"] = sample_note

    return briefing


def build_all_briefings(
    findings_data: dict | None,
    pair_forensics_data: dict | None,
    source_data_dir: Path,
) -> dict[str, Any]:
    """Build briefings for all sheets and produce the artifact dict."""
    # Group findings by (workbook, sheet)
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)

    for f in (findings_data or {}).get("findings", []):
        wb = f.get("workbook", "")
        sh = f.get("sheet", "")
        if wb and sh:
            grouped[(wb, sh)].append(f)

    for f in (pair_forensics_data or {}).get("findings", []):
        wb = f.get("workbook", "")
        sh = f.get("sheet", "")
        if wb and sh:
            grouped[(wb, sh)].append(f)

    briefings: list[dict[str, Any]] = []
    for (wb, sh), findings in sorted(grouped.items()):
        briefing = build_sheet_briefing(wb, sh, findings, source_data_dir)
        briefings.append(briefing)

    return {
        "schema_version": "1.0",
        "created_by": "engine/static_audit/tools/source_data_sheet_briefing.py",
        "sheet_count": len(briefings),
        "sheets": briefings,
    }
